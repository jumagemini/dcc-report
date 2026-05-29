import os
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from celery.result import AsyncResult
from .tasks import extract_serial_numbers
import tempfile
from django.template.loader import render_to_string
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files import File
from django.urls import reverse, reverse_lazy
from django.shortcuts import render
from django.views.generic import UpdateView, DeleteView
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from weasyprint import HTML
from io import BytesIO
from .models import DCC, Institution, InstitutionPhoto, UserProfile, UserDCCLimit, DeletionRequest,Notification
from .forms import InstitutionForm, PhotoUploadForm
from .utils import generate_dcc_stickers_docx, generate_dcc_photos_docx, save_signature_from_data_url
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
import base64
import logging
import re
import zipfile

logger = logging.getLogger(__name__)

def user_allowed_dcc(user, dcc):
    profile = UserProfile.objects.filter(user=user).first()
    return profile and profile.allowed_dccs.filter(pk=dcc.pk).exists()

def _process_institution_form(request, dcc, template_name):
    """
    Handle both GET and POST for institution creation.
    template_name: which template to use for rendering (e.g. 'reports/institution_form.html'
                   or 'reports/institution_active.html').
    """
    # ---------- Authorisation & limit check (already done by the caller) ----------
    # (We assume the caller already verified user_allowed_dcc and limit.)

    if request.method == 'POST':
        form = InstitutionForm(request.POST)
        photo_form = PhotoUploadForm(request.POST, request.FILES)

        if form.is_valid() and photo_form.is_valid():
            # -----------------------------------------------------------------
            # Cross‑check: prevent photo upload without serial number
            # -----------------------------------------------------------------
            errors_found = False
            photo_to_serial = {
                'before_onu': ('onu_serial', 'ONU'),
                'before_ap1': ('indoor_ap1_serial', 'INDOOR AP1'),
                'before_ap2': ('indoor_ap2_serial', 'INDOOR AP2'),
                'before_ap3': ('indoor_ap3_serial', 'INDOOR AP3'),
                'before_out': ('outdoor_ap_serial', 'OUTDOOR AP1'),
                'after_onu': ('onu_serial', 'ONU'),
                'after_ap1': ('indoor_ap1_serial', 'INDOOR AP1'),
                'after_ap2': ('indoor_ap2_serial', 'INDOOR AP2'),
                'after_ap3': ('indoor_ap3_serial', 'INDOOR AP3'),
                'after_out': ('outdoor_ap_serial', 'OUTDOOR AP1'),
            }
            for photo_field, (serial_field, device_name) in photo_to_serial.items():
                if photo_form.cleaned_data.get(photo_field):
                    if not form.cleaned_data.get(serial_field):
                        photo_form.add_error(
                            photo_field,
                            f"Serial Number is required for {device_name} before uploading a photo."
                        )
                        errors_found = True

            if not errors_found:
                # Save the institution
                institution = form.save(commit=False)
                institution.dcc = dcc
                institution.created_by = request.user
                institution.save()
                signature_data = request.POST.get('signature_image', '')
                if signature_data:
                    save_signature_from_data_url(signature_data, institution)
                else:
                    # No new signature drawn – use user's profile default if institution has none
                    if not institution.signature and request.user.userprofile.signature:
                        institution.signature.save(
                            'signature.png',
                            File(open(request.user.userprofile.signature.path, 'rb')),
                            save=False
                        )
                institution.save()

                # Save photos
                device_mapping = {
                    'before_onu': ('before', 'ONU'),
                    'before_ap1': ('before', 'AP1'),
                    'before_ap2': ('before', 'AP2'),
                    'before_ap3': ('before', 'AP3'),
                    'before_out': ('before', 'OUT'),
                    'after_onu': ('after', 'ONU'),
                    'after_ap1': ('after', 'AP1'),
                    'after_ap2': ('after', 'AP2'),
                    'after_ap3': ('after', 'AP3'),
                    'after_out': ('after', 'OUT'),
                }
                for field_name, (photo_type, device_type) in device_mapping.items():
                    image_file = photo_form.cleaned_data.get(field_name)
                    if image_file:
                        InstitutionPhoto.objects.update_or_create(
                            institution=institution,
                            photo_type=photo_type,
                            device_type=device_type,
                            defaults={'image': image_file}
                        )

                pdf_url = reverse('institution_pdf_preview', args=[institution.pk])
                full_pdf_url = request.build_absolute_uri(pdf_url)

                if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'Installation for {institution.name} saved.',
                        'pdf_url': full_pdf_url
                    })
                messages.success(request, f'Installation for {institution.name} saved.')
                return redirect('institution_pdf', pk=institution.pk)

            # If cross‑validation errors exist, re‑render the form with errors
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                errors = {}
                errors.update(form.errors)
                errors.update(photo_form.errors)
                return JsonResponse({'success': False, 'errors': errors}, status=400)
            # For normal POST, fall through to render with errors
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                errors = {}
                errors.update(form.errors)
                errors.update(photo_form.errors)
                return JsonResponse({'success': False, 'errors': errors}, status=400)
    else:
        form = InstitutionForm()
        photo_form = PhotoUploadForm()

    return render(request, template_name, {
        'form': form,
        'photo_form': photo_form,
        'dcc': dcc,
        'existing_info': {},   # empty for create
    })

class InstitutionUpdateView(UpdateView):
    model = Institution
    form_class = InstitutionForm
    template_name = 'reports/institution_form.html'
    success_url = reverse_lazy('dashboard')  # or back to list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['dcc'] = self.object.dcc          # the existing DCC
        context['photo_form'] = PhotoUploadForm()
    
        # Build a dict of strings -> {'url': ..., 'id': ...} for the template
        existing_info = {}
        for photo in self.object.photos.all():
            key = f"{photo.photo_type}_{photo.device_type}"
            existing_info[key] = {
                'url': photo.image.url if photo.image else '',
                'id': photo.id,
            }
        context['existing_info'] = existing_info
        return context
    
    def form_valid(self, form):
        institution = self.object
        photo_form = PhotoUploadForm(self.request.POST, self.request.FILES)

        if not photo_form.is_valid():
            # Re‑render with photo_form errors
            context = self.get_context_data(form=form)
            context['photo_form'] = photo_form
            return self.render_to_response(context)

        # -----------------------------------------------------------------
        # Cross‑check: prevent photo upload without serial number
        # -----------------------------------------------------------------
        errors_found = False
        photo_to_serial = {
            'before_onu': ('onu_serial', 'ONU'),
            'before_ap1': ('indoor_ap1_serial', 'INDOOR AP1'),
            'before_ap2': ('indoor_ap2_serial', 'INDOOR AP2'),
            'before_ap3': ('indoor_ap3_serial', 'INDOOR AP3'),
            'before_out': ('outdoor_ap_serial', 'OUTDOOR AP1'),
            'after_onu': ('onu_serial', 'ONU'),
            'after_ap1': ('indoor_ap1_serial', 'INDOOR AP1'),
            'after_ap2': ('indoor_ap2_serial', 'INDOOR AP2'),
            'after_ap3': ('indoor_ap3_serial', 'INDOOR AP3'),
            'after_out': ('outdoor_ap_serial', 'OUTDOOR AP1'),
        }
        for photo_field, (serial_field, device_name) in photo_to_serial.items():
            if photo_form.cleaned_data.get(photo_field):
                if not getattr(institution, serial_field, None):
                    photo_form.add_error(
                        photo_field,
                        f"Serial Number is required for {device_name} before uploading a photo."
                    )
                    errors_found = True

        if errors_found:
            context = self.get_context_data(form=form)
            context['photo_form'] = photo_form
            return self.render_to_response(context)
        # -----------------------------------------------------------------

        # Process photo removals
        for key, value in self.request.POST.items():
            if key.startswith('delete_photo_') and value:
                InstitutionPhoto.objects.filter(
                    id=value, institution=institution
                ).delete()

        # Process new uploads
        device_mapping = {
            'before_onu': ('before', 'ONU'),
            'before_ap1': ('before', 'AP1'),
            'before_ap2': ('before', 'AP2'),
            'before_ap3': ('before', 'AP3'),
            'before_out': ('before', 'OUT'),
            'after_onu': ('after', 'ONU'),
            'after_ap1': ('after', 'AP1'),
            'after_ap2': ('after', 'AP2'),
            'after_ap3': ('after', 'AP3'),
            'after_out': ('after', 'OUT'),
        }
        for field_name, (photo_type, device_type) in device_mapping.items():
            image_file = photo_form.cleaned_data.get(field_name)
            if image_file:
                InstitutionPhoto.objects.update_or_create(
                    institution=institution,
                    photo_type=photo_type,
                    device_type=device_type,
                    defaults={'image': image_file}
                )
        # Save the institution form (returns a redirect response)
        response = super().form_valid(form)

        # --- Save the digital signature ---
        signature_data = self.request.POST.get('signature_image', '')
        if signature_data:
            save_signature_from_data_url(signature_data, self.object)
        else:
            # Use user's default signature if institution still has none
            if not self.object.signature:
                user_profile = getattr(self.request.user, 'userprofile', None)
                if user_profile and user_profile.signature:
                    with open(user_profile.signature.path, 'rb') as f:
                        self.object.signature.save('signature.png', File(f), save=False)
        self.object.save()

        return response
    
    def dispatch(self, request, *args, **kwargs):
        if not user_allowed_dcc(request.user, self.get_object().dcc):
            return HttpResponseForbidden("You are not authorised for this DCC.")
        return super().dispatch(request, *args, **kwargs) 

class InstitutionDeleteView(DeleteView):
    model = Institution
    template_name = 'reports/institution_confirm_delete.html'
    success_url = reverse_lazy('dashboard')

    def dispatch(self, request, *args, **kwargs):
        # Obtain the institution without fully loading the object twice
        institution = self.get_object()
        if not user_allowed_dcc(request.user, institution.dcc):
            return HttpResponseForbidden("You are not authorised for this DCC.")
        return super().dispatch(request, *args, **kwargs)

@login_required
def dashboard(request):
    profile = UserProfile.objects.filter(user=request.user).first()
    if not profile:
        dccs = []   # or redirect to a "no access" page
    else:
        dccs = list(profile.allowed_dccs.all().order_by('name'))
        for dcc in dccs:
            # Attach user-specific limit info
            try:
                limit_obj = UserDCCLimit.objects.get(user=request.user, dcc=dcc)
                dcc.user_max = limit_obj.max_institutions  # None means unlimited
            except UserDCCLimit.DoesNotExist:
                dcc.user_max = None   # no limit set -> unlimited
            # current institutions created by this user in this DCC
            dcc.user_count = Institution.objects.filter(dcc=dcc, created_by=request.user).count()
    return render(request, 'reports/dashboard.html', {'dccs': dccs})

@login_required
def institution_list(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    # Optional: check user is allowed
    institutions = dcc.institutions.all().order_by('name')
    return render(request, 'reports/institution_list.html', {
        'dcc': dcc,
        'institutions': institutions
    })

def sanitize_filename(name):
    # Replace spaces with underscores and remove non-alphanumeric/underscore/hyphen
    name = name.replace(' ', '_')
    return re.sub(r'[^\w\-]', '', name)

def institution_create(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")

    # Limit check
    try:
        limit = UserDCCLimit.objects.get(user=request.user, dcc=dcc)
    except UserDCCLimit.DoesNotExist:
        limit = None
    if limit and limit.max_institutions is not None:
        count = Institution.objects.filter(dcc=dcc, created_by=request.user).count()
        if count >= limit.max_institutions:
            messages.warning(
                request,
                f"You have reached the maximum number of institutions ({limit.max_institutions}) for {dcc.name}."
            )
            return redirect('dashboard')

    return _process_institution_form(request, dcc, 'reports/institution_form.html')

def institution_create_active(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")

    # Limit check
    try:
        limit = UserDCCLimit.objects.get(user=request.user, dcc=dcc)
    except UserDCCLimit.DoesNotExist:
        limit = None
    if limit and limit.max_institutions is not None:
        count = Institution.objects.filter(dcc=dcc, created_by=request.user).count()
        if count >= limit.max_institutions:
            messages.warning(
                request,
                f"You have reached the maximum number of institutions ({limit.max_institutions}) for {dcc.name}."
            )
            return redirect('dashboard')

    return _process_institution_form(request, dcc, 'reports/institution_active.html')

def generate_dcc_excel(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    # In each view, before processing:
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")
    institutions = dcc.institutions.all().order_by('name')

    wb = openpyxl.Workbook()
    # ===================== SHEET 1: DEVICE SUMMARY =====================
    ws1 = wb.active
    ws1.title = "Device Summary"

    # --- Title Rows (with merged cells) ---
    # Row 2: Project name
    cell = ws1.cell(row=2, column=2, value=f"Project name: {dcc.project_name}")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=2, start_column=2, end_row=2, end_column=12)

    # Row 3: DCC name
    cell = ws1.cell(row=3, column=2, value=f"DCC: {dcc.name}")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=12)

    # Row 4: "Devices serial numbers"
    cell = ws1.cell(row=4, column=2, value="Devices serial numbers")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=4, start_column=2, end_row=4, end_column=12)

    # Row 5: DCC name again
    cell = ws1.cell(row=5, column=2, value=dcc.name)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=5, start_column=2, end_row=5, end_column=12)

    # Row 6: Institution count and group headers
    cell = ws1.cell(row=6, column=2, value=f"{institutions.count()} Institutions.")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=6, start_column=2, end_row=6, end_column=2)

    cell = ws1.cell(row=6, column=3, value="Indoor Access point")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=3, end_row=6, end_column=8)

    cell = ws1.cell(row=6, column=9, value="Outdoor AP")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=9, end_row=6, end_column=10)

    cell = ws1.cell(row=6, column=11, value="4 port ONU")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=11, end_row=6, end_column=12)

    # Row 7: Column headers (exactly as in file)
    headers = [
        "",  # A column will have index numbers
        "Name of Institution",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=7, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # --- Data Rows ---
    start_row = 8
    for idx, inst in enumerate(institutions, start=1):
        row = start_row + idx - 1
        # Index number in column A
        ws1.cell(row=row, column=1, value=idx)
        # Institution name column B
        ws1.cell(row=row, column=2, value=inst.name)

        # Indoor AP1 serial/location (columns C,D) - mandatory
        ws1.cell(row=row, column=3, value=inst.indoor_ap1_serial)
        ws1.cell(row=row, column=4, value=inst.indoor_ap1_location)

        # Indoor AP2 serial/location (columns E,F) - optional, replace empty with 'N/A'
        ap2_serial = inst.indoor_ap2_serial if inst.indoor_ap2_serial else 'N/A'
        ap2_location = inst.indoor_ap2_location if inst.indoor_ap2_location else 'N/A'
        cell_e = ws1.cell(row=row, column=5, value=ap2_serial)
        cell_f = ws1.cell(row=row, column=6, value=ap2_location)
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_f.alignment = Alignment(horizontal='center', vertical='center')

        # Indoor AP3 serial/location (columns G,H) - optional, replace empty with 'N/A'
        ap3_serial = inst.indoor_ap3_serial if inst.indoor_ap3_serial else 'N/A'
        ap3_location = inst.indoor_ap3_location if inst.indoor_ap3_location else 'N/A'
        cell_g = ws1.cell(row=row, column=7, value=ap3_serial)
        cell_h = ws1.cell(row=row, column=8, value=ap3_location)
        cell_g.alignment = Alignment(horizontal='center', vertical='center')
        cell_h.alignment = Alignment(horizontal='center', vertical='center')

        # Outdoor AP serial/location (columns I,J) - mandatory
        ws1.cell(row=row, column=9, value=inst.outdoor_ap_serial)
        ws1.cell(row=row, column=10, value=inst.outdoor_ap_location)

        # ONU serial/location (columns K,L) - mandatory
        ws1.cell(row=row, column=11, value=inst.onu_serial)
        ws1.cell(row=row, column=12, value=inst.onu_location)

    # --- Column Widths ---
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 45
    for col in ['C','E','G','I','K']:
        ws1.column_dimensions[col].width = 22
    for col in ['D','F','H','J','L']:
        ws1.column_dimensions[col].width = 20

        # ===================== SHEET 2: PHOTOS =====================
    ws2 = wb.create_sheet(title="Installation Photos")

    # Set column widths
    col_widths = {
        'A': 35, 'B': 5, 'C': 30, 'D': 5, 'E': 5,
        'F': 30, 'G': 5, 'H': 5, 'I': 30, 'J': 5,
        'K': 5, 'L': 30, 'M': 30
    }
    for col, width in col_widths.items():
        ws2.column_dimensions[col].width = width

    device_cols = {
        'ONU': 1, 'AP1': 3, 'AP2': 6, 'AP3': 9, 'OUT': 12,
    }

    IMG_WIDTH = 200
    IMG_HEIGHT = 150
    IMAGE_ROW_HEIGHT = 160
    LABEL_ROW_HEIGHT = 25   # height for label rows

    current_row = 1
    for idx, inst in enumerate(institutions, start=1):
        # Institution Name (merged)
        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
        cell = ws2.cell(row=current_row, column=1, value=f"{idx}. {inst.name.upper()}")
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='left')
        current_row += 1

        # Blank row
        current_row += 1

        # BEFORE INSTALLATION header
        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
        cell = ws2.cell(row=current_row, column=1, value="BEFORE INSTALLATION")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left')
        current_row += 1

        # Image row for BEFORE
        before_image_row = current_row
        ws2.row_dimensions[before_image_row].height = IMAGE_ROW_HEIGHT
        current_row += 1

        # Place before images without offset (anchor to top-left of cell)
        for dev, col_idx in device_cols.items():
            photo = inst.photos.filter(photo_type='before', device_type=dev).first()
            cell_ref = f'{get_column_letter(col_idx)}{before_image_row}'
            if photo:
                img = XLImage(photo.image.path)
                img.width = IMG_WIDTH
                img.height = IMG_HEIGHT
                ws2.add_image(img, cell_ref)
            else:
                cell = ws2.cell(row=before_image_row, column=col_idx, value='N/A')
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')    

        # Label row for BEFORE
        label_row = current_row
        ws2.row_dimensions[label_row].height = LABEL_ROW_HEIGHT
        current_row += 1

        # Determine installed devices
        has_device = {
            'ONU': bool(inst.onu_serial),
            'AP1': bool(inst.indoor_ap1_serial),
            'AP2': bool(inst.indoor_ap2_serial),
            'AP3': bool(inst.indoor_ap3_serial),
            'OUT': bool(inst.outdoor_ap_serial),
        }

        labels = {}
        if has_device['ONU']:
            labels['ONU'] = f"ONU {inst.onu_location}" if inst.onu_location else "ONU"
        if has_device['AP1']:
            labels['AP1'] = f"INDOOR AP1 {inst.indoor_ap1_location}" if inst.indoor_ap1_location else "INDOOR AP1"
        if has_device['AP2']:
            labels['AP2'] = f"INDOOR AP2 {inst.indoor_ap2_location}" if inst.indoor_ap2_location else "INDOOR AP2"
        if has_device['AP3']:
            labels['AP3'] = f"INDOOR AP3 {inst.indoor_ap3_location}" if inst.indoor_ap3_location else "INDOOR AP3"
        if has_device['OUT']:
            labels['OUT'] = f"OUTDOOR AP1 {inst.outdoor_ap_location}" if inst.outdoor_ap_location else "OUTDOOR AP1"

        # Write BEFORE labels
        if 'ONU' in labels:
            cell = ws2.cell(row=label_row, column=1, value=labels['ONU'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP1' in labels:
            ws2.merge_cells(start_row=label_row, start_column=3, end_row=label_row, end_column=4)
            cell = ws2.cell(row=label_row, column=3, value=labels['AP1'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP2' in labels:
            ws2.merge_cells(start_row=label_row, start_column=6, end_row=label_row, end_column=7)
            cell = ws2.cell(row=label_row, column=6, value=labels['AP2'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP3' in labels:
            ws2.merge_cells(start_row=label_row, start_column=9, end_row=label_row, end_column=10)
            cell = ws2.cell(row=label_row, column=9, value=labels['AP3'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'OUT' in labels:
            ws2.merge_cells(start_row=label_row, start_column=12, end_row=label_row, end_column=13)
            cell = ws2.cell(row=label_row, column=12, value=labels['OUT'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # AFTER INSTALLATION header
        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=13)
        cell = ws2.cell(row=current_row, column=1, value="AFTER INSTALLATION")
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='left')
        current_row += 1

        # Image row for AFTER
        after_image_row = current_row
        ws2.row_dimensions[after_image_row].height = IMAGE_ROW_HEIGHT
        current_row += 1

        # Place after images
        for dev, col_idx in device_cols.items():
            photo = inst.photos.filter(photo_type='after', device_type=dev).first()
            cell_ref = f'{get_column_letter(col_idx)}{after_image_row}'
            if photo:
                img = XLImage(photo.image.path)
                img.width = IMG_WIDTH
                img.height = IMG_HEIGHT
                ws2.add_image(img, cell_ref)
            else:
                cell = ws2.cell(row=after_image_row, column=col_idx, value='N/A')
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')


        # Label row for AFTER
        label_row = current_row
        ws2.row_dimensions[label_row].height = LABEL_ROW_HEIGHT
        current_row += 1

        # Write AFTER labels (same as BEFORE)
        if 'ONU' in labels:
            cell = ws2.cell(row=label_row, column=1, value=labels['ONU'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP1' in labels:
            ws2.merge_cells(start_row=label_row, start_column=3, end_row=label_row, end_column=4)
            cell = ws2.cell(row=label_row, column=3, value=labels['AP1'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP2' in labels:
            ws2.merge_cells(start_row=label_row, start_column=6, end_row=label_row, end_column=7)
            cell = ws2.cell(row=label_row, column=6, value=labels['AP2'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'AP3' in labels:
            ws2.merge_cells(start_row=label_row, start_column=9, end_row=label_row, end_column=10)
            cell = ws2.cell(row=label_row, column=9, value=labels['AP3'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        if 'OUT' in labels:
            ws2.merge_cells(start_row=label_row, start_column=12, end_row=label_row, end_column=13)
            cell = ws2.cell(row=label_row, column=12, value=labels['OUT'])
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Spacer between institutions
        current_row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{dcc.name}_report.xlsx"'
    return response

def get_image_base64(filename):
    image_path = settings.BASE_DIR / 'static' / 'images' / filename
    try:
        with open(image_path, 'rb') as f:
            return base64.b64encode(f.read()).decode('utf-8')
    except FileNotFoundError:
        logger.error(f"Logo file not found: {image_path}")
        return ""  # return empty string or a placeholder base64
    
def get_signature_base64(institution, user=None):
    if institution.signature:
        path = institution.signature.path
    elif user and user.userprofile.signature:
        path = user.userprofile.signature.path
    else:
        return ''

    try:
        with open(path, 'rb') as f:
            return base64.b64encode(f.read()).decode('utf-8')
    except Exception:
        return ''   

def preview_institution_pdf(request, pk):
    institution = get_object_or_404(Institution, pk=pk)
    # In each view, before processing:
    dcc = institution.dcc
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")
    print(f"DEBUG: project_no = '{institution.project_no}'")  # check console

    # Build sequential device list
    devices = []
    item_no = 1
    
    # ONU (mandatory)
    devices.append({
        'item_no': item_no,
        'name': 'ONU',
        'serial': institution.onu_serial,
        'quantity': 1,
        'location': institution.onu_location,
    })
    item_no += 1
    
    # Indoor AP1 (mandatory)
    devices.append({
        'item_no': item_no,
        'name': 'INDOOR AP1',
        'serial': institution.indoor_ap1_serial,
        'quantity': 1,
        'location': institution.indoor_ap1_location,
    })
    item_no += 1
    
    # Indoor AP2 (optional)
    if institution.indoor_ap2_serial:
        devices.append({
            'item_no': item_no,
            'name': 'INDOOR AP2',
            'serial': institution.indoor_ap2_serial,
            'quantity': 1,
            'location': institution.indoor_ap2_location,
        })
        item_no += 1
    
    # Indoor AP3 (optional)
    if institution.indoor_ap3_serial:
        devices.append({
            'item_no': item_no,
            'name': 'INDOOR AP3',
            'serial': institution.indoor_ap3_serial,
            'quantity': 1,
            'location': institution.indoor_ap3_location,
        })
        item_no += 1
    
    # Outdoor AP1 (mandatory)
    devices.append({
        'item_no': item_no,
        'name': 'OUTDOOR AP1',
        'serial': institution.outdoor_ap_serial,
        'quantity': 1,
        'location': institution.outdoor_ap_location,
    })

    try:
        context = {
            'institution': institution,
            'devices': devices,
            'logo_left': get_image_base64('kplc.png'),
            'logo_right': get_image_base64('ict.png'),
            'signature_base64': get_signature_base64(institution, institution.created_by),  # ← new
        }
        html_string = render_to_string('reports/institution_pdf.html', context)
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        # Create Safe Filename
        dcc_name = sanitize_filename(institution.dcc.name)
        inst_name = sanitize_filename(institution.name)
        filename = f"{dcc_name}_{inst_name}.pdf"

        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response
    except Exception as e:
        logger.exception("PDF generation failed")
        return HttpResponse(f"Error generating PDF: {e}", status=500)

def generate_institution_pdf(request, pk):
    institution = get_object_or_404(Institution, pk=pk)
    # In each view, before processing:
    dcc = institution.dcc
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")
    # Build sequential device list
    devices = []
    item_no = 1

    # ONU
    devices.append({
        'item_no': item_no,
        'name': 'ONU',
        'serial': institution.onu_serial,
        'quantity': 1,
        'location': institution.onu_location,
    })
    item_no += 1

    # Indoor AP1
    devices.append({
        'item_no': item_no,
        'name': 'INDOOR AP1',
        'serial': institution.indoor_ap1_serial,
        'quantity': 1,
        'location': institution.indoor_ap1_location,
    })
    item_no += 1

    # Indoor AP2 (optional)
    if institution.indoor_ap2_serial:
        devices.append({
            'item_no': item_no,
            'name': 'INDOOR AP2',
            'serial': institution.indoor_ap2_serial,
            'quantity': 1,
            'location': institution.indoor_ap2_location,
        })
        item_no += 1

    # Indoor AP3 (optional)
    if institution.indoor_ap3_serial:
        devices.append({
            'item_no': item_no,
            'name': 'INDOOR AP3',
            'serial': institution.indoor_ap3_serial,
            'quantity': 1,
            'location': institution.indoor_ap3_location,
        })
        item_no += 1

    # Outdoor AP1
    devices.append({
        'item_no': item_no,
        'name': 'OUTDOOR AP1',
        'serial': institution.outdoor_ap_serial,
        'quantity': 1,
        'location': institution.outdoor_ap_location,
    })

    context = {
        'institution': institution,
        'devices': devices,
        'logo_left': get_image_base64('kplc.png'),
        'logo_right': get_image_base64('ict.png'),
        'signature_base64': get_signature_base64(institution, institution.created_by),  # ← new
    }
    html_string = render_to_string('reports/institution_pdf.html', context)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')

    dcc_name = sanitize_filename(institution.dcc.name)
    inst_name = sanitize_filename(institution.name)
    filename = f"{dcc_name}_{inst_name}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response

def download_dcc_stickers(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    # In each view, before processing:
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")
    buffer = generate_dcc_stickers_docx(dcc)
    filename = f"{dcc.name.replace(' ', '_')}_Device_Labels.docx"
    response = FileResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def download_photo_bucket(request, pk):
    institution = get_object_or_404(Institution, pk=pk)
    dcc = institution.dcc
    # Check authorisation
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")

    # Mapping of device_type -> (label, serial)
    device_info = {
        'ONU':   ('ONU',              institution.onu_serial),
        'AP1':   ('INDOOR_AP1',       institution.indoor_ap1_serial),
        'AP2':   ('INDOOR_AP2',       institution.indoor_ap2_serial or ''),
        'AP3':   ('INDOOR_AP3',       institution.indoor_ap3_serial or ''),
        'OUT':   ('OUTDOOR_AP1',      institution.outdoor_ap_serial),
    }

    # Prepare safe institution name
    inst_name = sanitize_filename(institution.name)

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for photo in institution.photos.all():
            # Determine folder
            if photo.photo_type == 'before':
                folder = f"{inst_name}_before_installation_photos"
            else:
                folder = f"{inst_name}_after_installation_photos"

            # Build a descriptive filename
            label, serial = device_info.get(photo.device_type, (photo.device_type, ''))
            original_ext = os.path.splitext(photo.image.path)[1]  # e.g., '.jpg'

            if photo.photo_type == 'before' and serial:
                new_filename = f"{inst_name}_{label}_{serial}{original_ext}"
            else:
                new_filename = f"{inst_name}_{label}{original_ext}"

            zip_path = os.path.join(folder, new_filename)

            # Write the photo content
            try:
                with open(photo.image.path, 'rb') as f:
                    zipf.writestr(zip_path, f.read())
            except FileNotFoundError:
                logger.warning(f"Photo file not found: {photo.image.path}")

    buffer.seek(0)

    # Response
    dcc_name = sanitize_filename(dcc.name)
    zip_filename = f"{dcc_name}_{inst_name}.zip"

    response = HttpResponse(buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response

def download_dcc_photos(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")

    institutions = dcc.institutions.all()
    master_buffer = BytesIO()

    with zipfile.ZipFile(master_buffer, 'w', zipfile.ZIP_DEFLATED) as master_zip:
        for institution in institutions:
            # Create an empty "institution bucket" placeholder folder
            dcc_name = sanitize_filename(dcc.name)
            inst_name = sanitize_filename(institution.name)
            base_folder = f"{dcc_name}_{inst_name}"   # folder per institution

            device_info = {
                'ONU':   ('ONU',              institution.onu_serial),
                'AP1':   ('INDOOR_AP1',       institution.indoor_ap1_serial),
                'AP2':   ('INDOOR_AP2',       institution.indoor_ap2_serial or ''),
                'AP3':   ('INDOOR_AP3',       institution.indoor_ap3_serial or ''),
                'OUT':   ('OUTDOOR_AP1',      institution.outdoor_ap_serial),
            }

            for photo in institution.photos.all():
                if photo.photo_type == 'before':
                    subfolder = f"{base_folder}/{inst_name}_before_installation_photos"
                else:
                    subfolder = f"{base_folder}/{inst_name}_after_installation_photos"

                label, serial = device_info.get(photo.device_type, (photo.device_type, ''))
                original_ext = os.path.splitext(photo.image.path)[1]

                if photo.photo_type == 'before' and serial:
                    new_filename = f"{inst_name}_{label}_{serial}{original_ext}"
                else:
                    new_filename = f"{inst_name}_{label}{original_ext}"

                zip_path = os.path.join(subfolder, new_filename)

                try:
                    with open(photo.image.path, 'rb') as f:
                        master_zip.writestr(zip_path, f.read())
                except FileNotFoundError:
                    logger.warning(f"File missing: {photo.image.path}")

    master_buffer.seek(0)
    zip_filename = f"{sanitize_filename(dcc.name)}_Photo_Dump.zip"

    response = HttpResponse(master_buffer.read(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response

@csrf_exempt   # only because we handle via fetch – you can also include CSRF token properly
def request_institution_deletion(request, pk):
    institution = get_object_or_404(Institution, pk=pk)
    dcc = institution.dcc
    if not user_allowed_dcc(request.user, dcc):
        return JsonResponse({'success': False, 'error': 'You are not authorised for this DCC.'}, status=403)

    if request.method == 'POST':
        # Check if there is already a pending request for this institution
        existing = DeletionRequest.objects.filter(institution=institution, status='pending').first()
        if existing:
            return JsonResponse({
                'success': True,
                'message': (
                    f"Deletion request for {dcc.name}_{institution.name} "
                    "already exists and is awaiting admin approval."
                )
            })

        DeletionRequest.objects.create(institution=institution, requested_by=request.user)
        return JsonResponse({
            'success': True,
            'message': (
                f"{dcc.name}_{institution.name} Deletion Request Received, "
                "Awaiting Admin Approval!"
            )
        })
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)

@login_required
def notifications(request):
    notifications_list = Notification.objects.filter(user=request.user).order_by('-created_at')
    # Mark all as read when viewing
    notifications_list.filter(is_read=False).update(is_read=True)
    return render(request, 'reports/notifications.html', {'notifications': notifications_list})

def download_dcc_photos_docx(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")
    buffer = generate_dcc_photos_docx(dcc)
    filename = f"{sanitize_filename(dcc.name)}_Photo_Report.docx"
    response = FileResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def download_sheet1_excel(request, dcc_id):
    dcc = get_object_or_404(DCC, pk=dcc_id)
    if not user_allowed_dcc(request.user, dcc):
        return HttpResponseForbidden("You are not authorised for this DCC.")

    institutions = dcc.institutions.all().order_by('name')

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Device Summary"

    # --- Title Rows (identical to generate_dcc_excel) ---
    cell = ws1.cell(row=2, column=2, value=f"Project name: {dcc.project_name}")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=2, start_column=2, end_row=2, end_column=12)

    cell = ws1.cell(row=3, column=2, value=f"DCC: {dcc.name}")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=12)

    cell = ws1.cell(row=4, column=2, value="Devices serial numbers")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=4, start_column=2, end_row=4, end_column=12)

    cell = ws1.cell(row=5, column=2, value=dcc.name)
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=5, start_column=2, end_row=5, end_column=12)

    cell = ws1.cell(row=6, column=2, value=f"{institutions.count()} Institutions.")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left')
    ws1.merge_cells(start_row=6, start_column=2, end_row=6, end_column=2)

    cell = ws1.cell(row=6, column=3, value="Indoor Access point")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=3, end_row=6, end_column=8)

    cell = ws1.cell(row=6, column=9, value="Outdoor AP")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=9, end_row=6, end_column=10)

    cell = ws1.cell(row=6, column=11, value="4 port ONU")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    ws1.merge_cells(start_row=6, start_column=11, end_row=6, end_column=12)

    headers = [
        "", "Name of Institution",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location",
        "Serial Numbers", "Location"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=7, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    start_row = 8
    for idx, inst in enumerate(institutions, start=1):
        row = start_row + idx - 1
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=inst.name)

        ws1.cell(row=row, column=3, value=inst.indoor_ap1_serial)
        ws1.cell(row=row, column=4, value=inst.indoor_ap1_location)

        ap2_serial = inst.indoor_ap2_serial if inst.indoor_ap2_serial else 'N/A'
        ap2_location = inst.indoor_ap2_location if inst.indoor_ap2_location else 'N/A'
        cell_e = ws1.cell(row=row, column=5, value=ap2_serial)
        cell_f = ws1.cell(row=row, column=6, value=ap2_location)
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_f.alignment = Alignment(horizontal='center', vertical='center')

        ap3_serial = inst.indoor_ap3_serial if inst.indoor_ap3_serial else 'N/A'
        ap3_location = inst.indoor_ap3_location if inst.indoor_ap3_location else 'N/A'
        cell_g = ws1.cell(row=row, column=7, value=ap3_serial)
        cell_h = ws1.cell(row=row, column=8, value=ap3_location)
        cell_g.alignment = Alignment(horizontal='center', vertical='center')
        cell_h.alignment = Alignment(horizontal='center', vertical='center')

        ws1.cell(row=row, column=9, value=inst.outdoor_ap_serial)
        ws1.cell(row=row, column=10, value=inst.outdoor_ap_location)
        ws1.cell(row=row, column=11, value=inst.onu_serial)
        ws1.cell(row=row, column=12, value=inst.onu_location)

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 45
    for col in ['C','E','G','I','K']:
        ws1.column_dimensions[col].width = 22
    for col in ['D','F','H','J','L']:
        ws1.column_dimensions[col].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{dcc.name}_Device_Summary.xlsx"'
    return response

# IMAGE S/N PROCESSING
@csrf_exempt
@require_http_methods(["POST"])
def ocr_upload(request):
    if 'image' not in request.FILES:
        return JsonResponse({'error': 'No image provided'}, status=400)

    image_file = request.FILES['image']
    # Preserve the original file extension
    ext = os.path.splitext(image_file.name)[1] or '.jpg'
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        for chunk in image_file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name

    device_type = request.POST.get('device_type', 'AP1')
    task = extract_serial_numbers.delay(tmp_path, device_type)
    return JsonResponse({'task_id': task.id, 'device_type': device_type})


@csrf_exempt
@require_http_methods(["GET"])
def ocr_status(request, task_id):
    """Check task status and return results."""
    task = AsyncResult(task_id)
    if task.state == 'PENDING':
        return JsonResponse({'status': 'pending'})
    elif task.state == 'SUCCESS':
        # Remove the temporary file after successful processing
        # (the task doesn't delete it – you could add that logic in the task)
        return JsonResponse({'status': 'complete', 'serials': task.result})
    elif task.state == 'FAILURE':
        return JsonResponse({'status': 'failed', 'error': str(task.info)})
    else:
        return JsonResponse({'status': task.state})